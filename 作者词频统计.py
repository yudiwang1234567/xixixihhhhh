# 提取信息并计数
import openpyxl
wb=openpyxl.workbook()
wb=openpyxl.load_workbook('作者')
sh=wb['Sheet1']
row=sh.max_row  #行数
cons=[] #将所有数据append成一个list

for cases in list(sh.row)(1:15):
    case_D=cases[1].value  #关键词所在列
    cons.append(case_D)

result_new=[]
res2=[]

for con in cons:
    fenci=con.strip().split(';') #关键词以分号分割
    result_new.append(fenci)

new=sum(result_new,[])
res2.append(new)
#print(res2) #res2的形式是[‘aaa','bbb','ccc']

import collections
dic=collections.counter(res2[0])

from openpyxl import workbook
workbook=workbook()
i=2   #默认sheet
sh=workbook.active
sh.title='count' #sheet名
for key in dic:
    sh.cell(row=1,column=1,value='关键词')
    sh.cell(row=1,column=2,value='词频')
    sh.cell(row=1,column=1,value=str(key))
    sh.cell(row=i,column=2,value=dic[key])
    i+=1

workbook.save(r'name.xlsx')





